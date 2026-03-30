#!/usr/bin/env python3
"""
Attendance PDF to Excel Converter - FastAPI Backend
Converts Japanese attendance PDFs to Excel files
"""

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from calendar import monthrange
from datetime import datetime
import json
from pathlib import Path
import re
import sqlite3
from uuid import uuid4
import zipfile

# Excel and PDF libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pdfplumber

# Create app
app = FastAPI(title="Attendance PDF Converter", version="2.0")

BASE_DIR = Path(__file__).resolve().parent
EMPLOYEE_ROSTER_PATH = BASE_DIR / "employee_roster.json"
EMPLOYEE_ROSTER = json.loads(EMPLOYEE_ROSTER_PATH.read_text(encoding="utf-8"))
EMPLOYEE_ROSTER_BY_ID = {
    employee["employee_code"]: employee["name"]
    for employee in EMPLOYEE_ROSTER
}
DATABASE_PATH = BASE_DIR / "attendance.sqlite3"

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create upload directory
UPLOAD_DIR = Path("/tmp/attendance_uploads")
EXPORT_DIR = Path("/tmp/attendance_exports")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

uploaded_files = []
parsed_data = []
EMPTY_MARKERS = {"", "__:_", "__:__", "----", "------", "早退"}
TIME_PATTERN = re.compile(r"^(?:(当|翌)\s*)?(\d{1,2}):(\d{2})$")
MONTH_PATTERN = re.compile(r"(20\d{2})[.\-/](\d{1,2})")

def get_db_connection() -> sqlite3.Connection:
    """Open a SQLite connection for attendance persistence."""
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    return connection

def init_db() -> None:
    """Create the SQLite tables needed for uploads and attendance rows."""
    with get_db_connection() as connection:
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("""
            CREATE TABLE IF NOT EXISTS upload_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_filename TEXT NOT NULL,
                export_filename TEXT NOT NULL,
                records_processed INTEGER NOT NULL DEFAULT 0,
                file_count INTEGER NOT NULL DEFAULT 1,
                converted_at TEXT NOT NULL,
                month_year TEXT NOT NULL
            )
        """)
        connection.execute("""
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                source_filename TEXT NOT NULL,
                employee_code TEXT NOT NULL,
                full_name TEXT NOT NULL,
                commute_time TEXT DEFAULT '',
                leave_time TEXT DEFAULT '',
                working_hours TEXT DEFAULT '',
                month_year TEXT NOT NULL,
                converted_at TEXT NOT NULL,
                has_data INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(batch_id) REFERENCES upload_batches(id)
            )
        """)
        connection.execute("""
            CREATE INDEX IF NOT EXISTS idx_attendance_records_month
            ON attendance_records(month_year)
        """)
        connection.execute("""
            CREATE INDEX IF NOT EXISTS idx_attendance_records_employee
            ON attendance_records(employee_code, month_year)
        """)
        connection.execute("""
            CREATE INDEX IF NOT EXISTS idx_upload_batches_month
            ON upload_batches(month_year, converted_at)
        """)

init_db()

def clean_cell(value: object) -> str:
    """Normalize PDF table cell values for matching and display."""
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()

def is_employee_code(value: str) -> bool:
    """Attendance rows use 8-digit employee codes; subtotal rows do not."""
    return value.isdigit() and len(value) == 8

def find_column_index(header_row: list[str], keyword: str) -> int | None:
    """Find the first header cell containing the requested keyword."""
    for index, cell in enumerate(header_row):
        if keyword in cell:
            return index
    return None

def get_row_value(row: list[str], index: int | None) -> str:
    """Safely read a column value from an extracted table row."""
    if index is None or index >= len(row):
        return ""
    return row[index]

def normalize_blank_value(value: str) -> str:
    """Convert placeholder markers to blank strings."""
    return "" if value in EMPTY_MARKERS else value

def normalize_time_value(value: str, *, is_leave: bool = False) -> str:
    """
    Remove attendance prefixes and normalize late-night leave times.

    Leave times at or after midnight are represented in 24+ hour format,
    e.g. 1:00 AM becomes 25:00.
    """
    raw = clean_cell(value)
    if raw in EMPTY_MARKERS:
        return ""

    match = TIME_PATTERN.match(raw)
    if not match:
        return raw

    prefix, hour_str, minute_str = match.groups()
    hour = int(hour_str)
    minute = int(minute_str)

    if is_leave and (prefix == "翌" or hour <= 6):
        hour += 24

    return f"{hour}:{minute:02d}"

def time_to_minutes(value: str) -> int | None:
    """Convert a normalized H:MM string into absolute minutes."""
    normalized = normalize_blank_value(clean_cell(value))
    if not normalized:
        return None

    match = re.match(r"^(\d{1,2}):(\d{2})$", normalized)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    return hour * 60 + minute

def calculate_working_hours(start_time: str, leave_time: str) -> str:
    """Calculate one working-hours value like 8:00 hr from start/end times."""
    start_minutes = time_to_minutes(start_time)
    leave_minutes = time_to_minutes(leave_time)
    if start_minutes is None or leave_minutes is None or leave_minutes < start_minutes:
        return ""

    diff_minutes = leave_minutes - start_minutes
    hours = diff_minutes // 60
    minutes = diff_minutes % 60
    return f"{hours}:{minutes:02d} hr"

def parse_working_hours_minutes(value: str) -> int | None:
    """Convert values like 8:01 hr into minutes for dashboard calculations."""
    normalized = normalize_blank_value(clean_cell(value)).replace(" hr", "")
    if not normalized:
        return None

    match = re.match(r"^(\d+):(\d{2})$", normalized)
    if not match:
        return None

    return int(match.group(1)) * 60 + int(match.group(2))

def format_average_hours(minutes_values: list[int]) -> float:
    """Return an hours float rounded to 2 decimals for API summaries."""
    if not minutes_values:
        return 0.0
    return round((sum(minutes_values) / len(minutes_values)) / 60, 2)

def record_has_data(record: dict[str, str]) -> bool:
    """Check whether a roster row contains at least one non-blank value."""
    return any(record.get(field) for field in ("commute_time", "leave_time", "working_hours"))

def count_records_with_data(records: list[dict[str, str]]) -> int:
    """Count output rows that contain actual attendance data."""
    return sum(1 for record in records if record_has_data(record))

def extract_month_year(filename: str) -> str:
    """Derive a YYYY-MM value from the uploaded filename when available."""
    match = MONTH_PATTERN.search(filename)
    if not match:
        return datetime.now().strftime("%Y-%m")

    year, month = match.groups()
    return f"{year}-{int(month):02d}"

def resolve_month(month: str | None = None) -> str:
    """Choose the requested month or fall back to the latest saved batch."""
    if month:
        return month

    with get_db_connection() as connection:
        row = connection.execute(
            "SELECT month_year FROM upload_batches ORDER BY converted_at DESC LIMIT 1"
        ).fetchone()
    return row["month_year"] if row else datetime.now().strftime("%Y-%m")

def create_export_filename(original_name: str) -> str:
    """Build a safe Excel filename for an uploaded PDF."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).stem).strip("._")
    if not safe_stem:
        safe_stem = "attendance"
    return f"{safe_stem}_{timestamp}_{uuid4().hex[:8]}.xlsx"

def create_bundle_filename() -> str:
    """Build a zip filename for multi-file conversion."""
    return f"attendance_bundle_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.zip"

def save_batch_to_db(
    source_filename: str,
    export_filename: str,
    records: list[dict[str, str]],
    converted_at: str,
) -> tuple[int, int]:
    """Persist one converted upload batch and its attendance rows."""
    month_year = extract_month_year(source_filename)
    records_processed = count_records_with_data(records)

    with get_db_connection() as connection:
        existing_batch_ids = [
            row["id"]
            for row in connection.execute(
                """
                SELECT id
                FROM upload_batches
                WHERE source_filename = ? AND month_year = ?
                """,
                (source_filename, month_year),
            ).fetchall()
        ]
        if existing_batch_ids:
            connection.executemany(
                "DELETE FROM attendance_records WHERE batch_id = ?",
                [(batch_id,) for batch_id in existing_batch_ids],
            )
            connection.execute(
                """
                DELETE FROM upload_batches
                WHERE source_filename = ? AND month_year = ?
                """,
                (source_filename, month_year),
            )

        cursor = connection.execute(
            """
            INSERT INTO upload_batches (
                source_filename,
                export_filename,
                records_processed,
                file_count,
                converted_at,
                month_year
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (source_filename, export_filename, records_processed, 1, converted_at, month_year),
        )
        batch_id = cursor.lastrowid

        connection.executemany(
            """
            INSERT INTO attendance_records (
                batch_id,
                source_filename,
                employee_code,
                full_name,
                commute_time,
                leave_time,
                working_hours,
                month_year,
                converted_at,
                has_data
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    batch_id,
                    source_filename,
                    record["employee_code"],
                    record["name"],
                    record.get("commute_time", ""),
                    record.get("leave_time", ""),
                    record.get("working_hours", ""),
                    month_year,
                    converted_at,
                    1 if record_has_data(record) else 0,
                )
                for record in records
            ],
        )

    return batch_id, records_processed

def list_attendance_rows(month: str) -> list[sqlite3.Row]:
    """Fetch saved attendance rows for a given month."""
    with get_db_connection() as connection:
        return connection.execute(
            """
            SELECT employee_code, full_name, commute_time, leave_time, working_hours,
                   has_data, converted_at, source_filename
            FROM attendance_records
            WHERE month_year = ?
            ORDER BY employee_code, converted_at
            """,
            (month,),
        ).fetchall()

def build_preview_payload(
    filename: str,
    file_size: int,
    records: list[dict[str, str]],
    *,
    file_count: int = 1,
    preview_filename: str | None = None,
) -> dict[str, object]:
    """Create a consistent preview payload for one or many files."""
    return {
        "filename": filename,
        "preview_filename": preview_filename or filename,
        "file_size": file_size,
        "file_count": file_count,
        "status": "success",
        "records": records[:10],
        "total_records": len(records),
        "records_with_data": count_records_with_data(records),
        "extracted_at": datetime.now().isoformat(),
    }

async def parse_uploaded_pdf(file: UploadFile) -> tuple[str, int, list[dict[str, str]]]:
    """Save, parse, and roster-align an uploaded PDF."""
    original_name, file_path, content = await save_uploaded_pdf(file)
    records = apply_employee_roster(parse_pdf_data(file_path))
    return original_name, len(content), records

def export_records_to_excel(source_filename: str, records: list[dict[str, str]]) -> tuple[str, Path, int, int]:
    """Create one Excel file from roster-aligned records and save it to the DB."""
    excel_filename = create_export_filename(source_filename)
    excel_path = EXPORT_DIR / excel_filename
    create_excel_file(records, excel_path)

    converted_at = datetime.now().isoformat()
    batch_id, records_processed = save_batch_to_db(
        source_filename,
        excel_filename,
        records,
        converted_at,
    )

    uploaded_files.append({
        "source_filename": source_filename,
        "export_filename": excel_filename,
        "records_processed": records_processed,
        "converted_at": converted_at,
        "batch_id": batch_id,
    })

    return excel_filename, excel_path, records_processed, batch_id

def apply_employee_roster(records: list[dict[str, str]]) -> list[dict[str, str]]:
    """Return rows in the exact master roster order, with blanks for missing data."""
    records_by_id = {
        record["employee_code"]: record
        for record in records
        if record["employee_code"] in EMPLOYEE_ROSTER_BY_ID
    }

    rostered_records = []
    for employee in EMPLOYEE_ROSTER:
        code = employee["employee_code"]
        parsed_record = records_by_id.get(code, {})
        rostered_records.append({
            "employee_code": code,
            "name": employee["name"],
            "commute_time": parsed_record.get("commute_time", ""),
            "leave_time": parsed_record.get("leave_time", ""),
            "working_hours": parsed_record.get("working_hours", ""),
        })

    return rostered_records

def build_record_from_row(row: list[str], columns: dict[str, int | None]) -> dict[str, str] | None:
    """Convert one extracted PDF row into the API's record shape."""
    code = get_row_value(row, columns["employee_code"])
    if not is_employee_code(code):
        return None

    commute_time = normalize_time_value(get_row_value(row, columns["commute_time"]))
    leave_time = normalize_time_value(get_row_value(row, columns["leave_time"]), is_leave=True)

    return {
        "employee_code": code,
        "name": get_row_value(row, columns["name"]),
        "commute_time": commute_time,
        "leave_time": leave_time,
        "working_hours": calculate_working_hours(commute_time, leave_time),
    }

def parse_pdf_data(file_path: Path) -> list[dict[str, str]]:
    """Parse attendance PDF and extract data"""
    try:
        records = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        normalized_rows = [[clean_cell(cell) for cell in row] for row in table if row]
                        header_row = next((row for row in normalized_rows if any("個人" in cell for cell in row)), None)
                        if not header_row:
                            continue

                        columns = {
                            "employee_code": find_column_index(header_row, "個人"),
                            "name": find_column_index(header_row, "氏名"),
                            "commute_time": find_column_index(header_row, "出勤時刻"),
                            "leave_time": find_column_index(header_row, "退勤時刻"),
                        }

                        for row in normalized_rows:
                            record = build_record_from_row(row, columns)
                            if record:
                                records.append(record)

        if not records:
            raise ValueError("No attendance records were found in the PDF.")

        return records
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Unable to extract attendance data from the PDF: {exc}") from exc

def validate_pdf_content(content: bytes) -> None:
    """Reject empty uploads and obvious non-PDF files before parsing."""
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    if b"%PDF" not in content[:1024]:
        raise HTTPException(status_code=400, detail="Uploaded file is not a valid PDF.")

async def save_uploaded_pdf(file: UploadFile) -> tuple[str, Path, bytes]:
    """Persist the uploaded PDF using a safe, unique filename."""
    original_name = Path(file.filename or "").name
    if not original_name:
        raise HTTPException(status_code=400, detail="Uploaded file must include a filename.")

    if Path(original_name).suffix.lower() != ".pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are supported.")

    content = await file.read()
    validate_pdf_content(content)

    stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex}.pdf"
    file_path = UPLOAD_DIR / stored_name
    file_path.write_bytes(content)

    return original_name, file_path, content

def build_download_url(filename: str) -> str:
    """Return an app-local API path; the proxy prefix is added by the frontend."""
    return f"/api/download/{filename}"

def get_export_file_path(filename: str) -> Path:
    """Resolve a requested export filename inside the export directory only."""
    safe_name = Path(filename).name
    if safe_name != filename or Path(safe_name).suffix.lower() not in {".xlsx", ".zip"}:
        raise HTTPException(status_code=400, detail="Invalid filename.")

    file_path = (EXPORT_DIR / safe_name).resolve()
    if file_path.parent != EXPORT_DIR.resolve():
        raise HTTPException(status_code=400, detail="Invalid filename.")

    return file_path

def create_excel_file(records: list[dict[str, str]], filename: Path) -> Path:
    """Create readable Excel file with Japanese headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務表"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='MS ゴシック', size=11, bold=True, color="FFFFFF")
    data_font = Font(name='MS ゴシック', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add title
    ws['A1'] = f"勤務表　{datetime.now().strftime('%Y年%m月%d日')}"
    ws['A1'].font = Font(name='MS ゴシック', size=14, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # Add headers
    headers = ['個人コード', '氏名', '出勤時間', '退勤時間', '労働時間']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    ws.row_dimensions[3].height = 20
    
    # Add data rows
    for row_idx, record in enumerate(records, 4):
        ws.cell(row=row_idx, column=1).value = record.get('employee_code', '')
        ws.cell(row=row_idx, column=2).value = record.get('name', '')
        ws.cell(row=row_idx, column=3).value = record.get('commute_time', '')
        ws.cell(row=row_idx, column=4).value = record.get('leave_time', '')
        ws.cell(row=row_idx, column=5).value = record.get('working_hours', '')
        
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = border
        
        ws.row_dimensions[row_idx].height = 18
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    
    # Add summary row
    summary_row = len(records) + 5
    ws.cell(row=summary_row, column=1).value = "合計"
    ws.cell(row=summary_row, column=1).font = Font(name='MS ゴシック', size=10, bold=True)
    ws.cell(row=summary_row, column=2).value = f"従業員数: {len(records)}"
    ws.cell(row=summary_row, column=2).font = Font(name='MS ゴシック', size=10, bold=True)
    
    # Save file
    wb.save(filename)
    return filename

@app.get("/")
async def root():
    """Serve web interface"""
    return FileResponse(BASE_DIR / "index.html")

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    with get_db_connection() as connection:
        batch_count = connection.execute("SELECT COUNT(*) FROM upload_batches").fetchone()[0]
    return {
        "status": "healthy",
        "version": "2.0",
        "storage": "filesystem",
        "database": "sqlite",
        "saved_batches": batch_count,
        "timestamp": datetime.now().isoformat()
    }

@app.post("/api/preview")
async def preview_pdf(file: UploadFile = File(...)):
    """
    Upload PDF and preview extracted data
    """
    try:
        original_name, file_size, records = await parse_uploaded_pdf(file)
        preview_data = build_preview_payload(original_name, file_size, records)
        
        parsed_data.append(preview_data)
        return preview_data
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while previewing PDF: {exc}") from exc

@app.post("/api/preview-multiple")
async def preview_multiple_pdfs(files: list[UploadFile] = File(...)):
    """Preview a multi-file upload by showing the first file and total counts."""
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    try:
        parsed_files = [await parse_uploaded_pdf(file) for file in files]
        preview_name, preview_size, preview_records = parsed_files[0]
        total_records = sum(len(records) for _, _, records in parsed_files)
        records_with_data = sum(count_records_with_data(records) for _, _, records in parsed_files)

        preview_data = build_preview_payload(
            preview_name,
            preview_size,
            preview_records,
            file_count=len(parsed_files),
            preview_filename=preview_name,
        )
        preview_data["total_records"] = total_records
        preview_data["records_with_data"] = records_with_data
        preview_data["files"] = [
            {
                "filename": filename,
                "total_records": len(records),
                "records_with_data": count_records_with_data(records),
            }
            for filename, _file_size, records in parsed_files
        ]

        parsed_data.append(preview_data)
        return preview_data
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while previewing PDFs: {exc}") from exc

@app.post("/api/convert")
async def convert_pdf(file: UploadFile = File(...)):
    """
    Convert PDF to Excel
    Returns Excel file for download
    """
    try:
        original_name, _file_size, records = await parse_uploaded_pdf(file)
        excel_filename, _excel_path, records_processed, batch_id = export_records_to_excel(
            original_name,
            records,
        )

        return {
            "status": "success",
            "filename": excel_filename,
            "records_processed": records_processed,
            "roster_records": len(records),
            "batch_id": batch_id,
            "message": "PDFが正常に変換され、Excelファイルが生成されました",
            "download_url": build_download_url(excel_filename),
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while converting PDF: {exc}") from exc

@app.post("/api/convert-multiple")
async def convert_multiple_pdfs(files: list[UploadFile] = File(...)):
    """Convert multiple PDFs and return one zip bundle of Excel files."""
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    try:
        exports: list[tuple[str, Path, int]] = []
        total_records_processed = 0

        for file in files:
            original_name, _file_size, records = await parse_uploaded_pdf(file)
            excel_filename, excel_path, records_processed, _batch_id = export_records_to_excel(
                original_name,
                records,
            )
            exports.append((excel_filename, excel_path, len(records)))
            total_records_processed += records_processed

        bundle_filename = create_bundle_filename()
        bundle_path = EXPORT_DIR / bundle_filename
        with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for excel_filename, excel_path, _row_count in exports:
                archive.write(excel_path, arcname=excel_filename)

        return {
            "status": "success",
            "filename": bundle_filename,
            "file_count": len(exports),
            "records_processed": total_records_processed,
            "message": f"{len(exports)} PDF files were converted and bundled into one ZIP file.",
            "download_url": build_download_url(bundle_filename),
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while converting PDFs: {exc}") from exc

@app.get("/api/download/{filename}")
async def download_export(filename: str):
    """
    Download a generated Excel file or ZIP bundle.
    """
    file_path = get_export_file_path(filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    media_type = (
        "application/zip"
        if file_path.suffix.lower() == ".zip"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return FileResponse(
        file_path,
        media_type=media_type,
        filename=file_path.name
    )

@app.get("/api/dashboard/summary")
async def dashboard_summary(month: str | None = None):
    """
    Get dashboard summary statistics
    """
    resolved_month = resolve_month(month)
    rows = list_attendance_rows(resolved_month)
    active_rows = [row for row in rows if row["has_data"]]
    working_minutes = [
        minutes
        for minutes in (parse_working_hours_minutes(row["working_hours"]) for row in active_rows)
        if minutes is not None
    ]

    with get_db_connection() as connection:
        batch_row = connection.execute(
            """
            SELECT COUNT(*) AS batch_count, MAX(converted_at) AS latest_upload
            FROM upload_batches
            WHERE month_year = ?
            """,
            (resolved_month,),
        ).fetchone()

    year, month_value = resolved_month.split("-")
    return {
        "month": resolved_month,
        "total_employees": len({row["employee_code"] for row in active_rows}),
        "total_records": len(active_rows),
        "average_working_hours": format_average_hours(working_minutes),
        "days_in_month": monthrange(int(year), int(month_value))[1],
        "upload_batches": batch_row["batch_count"] if batch_row else 0,
        "latest_upload": batch_row["latest_upload"] if batch_row else None,
    }

@app.get("/api/dashboard/employees")
async def dashboard_employees(month: str | None = None):
    """
    Get all employees with summary
    """
    resolved_month = resolve_month(month)
    rows = list_attendance_rows(resolved_month)
    by_employee: dict[str, dict[str, object]] = {}

    for row in rows:
        if not row["has_data"]:
            continue

        code = row["employee_code"]
        summary = by_employee.setdefault(code, {
            "code": code,
            "name": row["full_name"],
            "days_worked": 0,
            "working_minutes": [],
        })
        summary["days_worked"] = int(summary["days_worked"]) + 1
        minutes = parse_working_hours_minutes(row["working_hours"])
        if minutes is not None:
            summary["working_minutes"].append(minutes)

    employees = []
    for code in [employee["employee_code"] for employee in EMPLOYEE_ROSTER]:
        summary = by_employee.get(code)
        if not summary:
            continue
        employees.append({
            "id": code,
            "code": code,
            "name": summary["name"],
            "days_worked": summary["days_worked"],
            "avg_working_hours": format_average_hours(summary["working_minutes"]),
        })

    return {
        "month": resolved_month,
        "employees": employees,
        "total_count": len(employees),
    }

@app.get("/api/dashboard/employee/{employee_code}")
async def dashboard_employee_detail(employee_code: str, month: str | None = None):
    """
    Get individual employee details
    """
    resolved_month = resolve_month(month)
    rows = [
        row for row in list_attendance_rows(resolved_month)
        if row["employee_code"] == employee_code
    ]
    working_minutes = [
        minutes
        for minutes in (parse_working_hours_minutes(row["working_hours"]) for row in rows)
        if minutes is not None
    ]
    late_count = sum(
        1
        for row in rows
        if row["has_data"] and (time_to_minutes(row["commute_time"]) or 0) > 9 * 60
    )
    early_count = sum(
        1
        for row in rows
        if row["has_data"] and row["leave_time"] and (time_to_minutes(row["leave_time"]) or 0) < 17 * 60
    )

    return {
        "employee_id": employee_code,
        "name": EMPLOYEE_ROSTER_BY_ID.get(employee_code, ""),
        "month": resolved_month,
        "records": [
            {
                "source_filename": row["source_filename"],
                "converted_at": row["converted_at"],
                "commute_time": row["commute_time"],
                "leave_time": row["leave_time"],
                "working_hours": row["working_hours"],
                "has_data": bool(row["has_data"]),
            }
            for row in rows
        ],
        "monthly_data": {
            "total_hours": round(sum(working_minutes) / 60, 2) if working_minutes else 0,
            "average_daily": format_average_hours(working_minutes),
            "late_count": late_count,
            "early_count": early_count,
        }
    }

@app.get("/api/dashboard/months")
async def dashboard_months():
    """
    Get available months with data
    """
    with get_db_connection() as connection:
        rows = connection.execute(
            "SELECT DISTINCT month_year FROM upload_batches ORDER BY month_year DESC"
        ).fetchall()

    months = [row["month_year"] for row in rows]
    return {
        "available_months": months,
        "latest_month": months[0] if months else None,
    }

# Serve static files
if (BASE_DIR / "static").exists():
    app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8002)
